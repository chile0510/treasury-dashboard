# -*- coding: utf-8 -*-
"""
sync_excel.py — Read Excel from OneDrive and update treasury_data.py.

Usage:
    python tools/sync_excel.py          # sync once
    python tools/sync_excel.py --push   # sync + git commit + push (auto-deploy)
"""

import os
import sys
import json
import subprocess
from datetime import datetime, date

# Force UTF-8 output on Windows (cp1252 can't handle Vietnamese)
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ONEDRIVE_BASE = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - Công Ty Cổ Phần Dịch Vụ Giao Hàng Nhanh",
)
EXCEL_DIR = os.path.join(
    ONEDRIVE_BASE,
    "Long Hoàng's files - Weekly Cash Report",
    "Working cap - Investment mapping",
)
EXCEL_FILENAME = "Mapping WC-Invest 26.04.xlsx"
EXCEL_PATH = os.path.join(EXCEL_DIR, EXCEL_FILENAME)

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(PROJECT_DIR, "lib", "treasury_data.py")

# ---------------------------------------------------------------------------
# Excel Column Mapping (1-indexed)
# ---------------------------------------------------------------------------
# Summary area (rows 2-7, left side)
COL_SUMMARY_LABEL = 3   # C
COL_SUMMARY_VALUE = 4   # D
COL_SUMMARY_YIELD = 5   # E
COL_SUMMARY_INTEREST = 6  # F

# Bank limit area (rows 2-7, right side)
COL_LIMIT_BANK = 8       # H
COL_LIMIT_DUNO = 9       # I
COL_LIMIT_HANMUC = 10    # J
COL_LIMIT_HANMUC2026 = 11  # K

# Loan transactions (row 9+, left side)
COL_LOAN_START = 3       # C
COL_LOAN_END = 4         # D
COL_LOAN_STATUS = 5      # E
COL_LOAN_BANK = 6        # F
COL_LOAN_AMOUNT = 7      # G
COL_LOAN_RATE = 8        # H

# Investment transactions (row 9+, right side)
COL_INV_START = 13       # M
COL_INV_END = 14         # N
COL_INV_STATUS = 15      # O
COL_INV_TYPE = 16        # P
COL_INV_BANK = 17        # Q
COL_INV_AMOUNT = 18      # R
COL_INV_RATE = 19        # S

DATA_START_ROW = 10  # First data row (after headers)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(ws, row, col):
    """Get cell value, return None if empty."""
    v = ws.cell(row=row, column=col).value
    return v


def _num(v, default=0):
    """Safely convert to float."""
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _date_str(v):
    """Convert datetime/date to 'YYYY-MM-DD' string."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).split(" ")[0]  # "2026-01-05 00:00:00" -> "2026-01-05"


def _status_flag(util_pct):
    """Determine limit control status from utilization %."""
    if util_pct >= 95:
        return "danger"
    elif util_pct >= 80:
        return "warning"
    return "safe"


# ---------------------------------------------------------------------------
# Excel Parsing
# ---------------------------------------------------------------------------

def parse_excel(filepath: str) -> dict:
    """Parse the Treasury Excel file and return FINANCIAL_DATA dict."""
    print(f"[SYNC] Reading: {filepath}")

    # Copy to temp file to avoid PermissionError when Excel has the file open
    import shutil
    import tempfile
    tmp_dir = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, "treasury_sync_temp.xlsx")
    try:
        shutil.copy2(filepath, tmp_path)
        print(f"[SYNC] Copied to temp: {tmp_path}")
    except PermissionError:
        print(f"[SYNC] WARNING: Cannot copy file (locked). Trying direct read...")
        tmp_path = filepath

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    ws = wb["Mapping"]

    # ---- 1. Limit Controls (rows 2-7, cols H-K) ----
    limit_controls = []
    for row in range(2, 8):
        bank = _cell(ws, row, COL_LIMIT_BANK)
        if not bank or str(bank).strip().lower() == "total":
            continue
        du_no = _num(_cell(ws, row, COL_LIMIT_DUNO))
        han_muc = _num(_cell(ws, row, COL_LIMIT_HANMUC))
        util = round(du_no / han_muc * 100, 1) if han_muc > 0 else 0
        room = han_muc - du_no

        limit_controls.append({
            "bank": str(bank).strip(),
            "duNo": du_no,
            "hanMuc": han_muc,
            "util": util,
            "room": max(room, 0),
            "status": _status_flag(util),
        })

    # ---- 2. Loans & Investments (row 10+) ----
    loans = []
    investments = []
    loan_invest_pairs = []  # For duration mismatch

    for row in range(DATA_START_ROW, ws.max_row + 1):
        # --- Loans (left side) ---
        loan_bank = _cell(ws, row, COL_LOAN_BANK)
        loan_status = str(_cell(ws, row, COL_LOAN_STATUS) or "").strip()
        loan_amount = _num(_cell(ws, row, COL_LOAN_AMOUNT))

        loan_entry = None
        if loan_bank and loan_amount > 0:
            loan_entry = {
                "bank": str(loan_bank).strip(),
                "amount": loan_amount,
                "rate": round(_num(_cell(ws, row, COL_LOAN_RATE)) * 100, 2),
                "startDate": _date_str(_cell(ws, row, COL_LOAN_START)),
                "endDate": _date_str(_cell(ws, row, COL_LOAN_END)),
                "status": loan_status if loan_status else "Outstanding",
            }
            if loan_status == "Outstanding":
                loans.append(loan_entry)

        # --- Investments (right side) ---
        inv_bank = _cell(ws, row, COL_INV_BANK)
        inv_status = str(_cell(ws, row, COL_INV_STATUS) or "").strip()
        inv_amount = _num(_cell(ws, row, COL_INV_AMOUNT))

        inv_entry = None
        if inv_bank and inv_amount > 0:
            inv_entry = {
                "bank": str(inv_bank).strip(),
                "amount": inv_amount,
                "rate": round(_num(_cell(ws, row, COL_INV_RATE)) * 100, 2),
                "type": str(_cell(ws, row, COL_INV_TYPE) or "").strip().upper(),
                "startDate": _date_str(_cell(ws, row, COL_INV_START)),
                "endDate": _date_str(_cell(ws, row, COL_INV_END)),
                "status": inv_status if inv_status else "Outstanding",
            }
            if inv_status == "Outstanding":
                investments.append(inv_entry)

        # --- Track pairs for duration mismatch ---
        if loan_entry and loan_status == "Outstanding" and inv_entry and inv_status == "Outstanding":
            loan_invest_pairs.append((loan_entry, inv_entry))

    # ---- 3. Summary (computed from filtered data) ----
    total_loan = sum(l["amount"] for l in loans)
    total_invest = sum(i["amount"] for i in investments)

    # Weighted average rates
    if total_loan > 0:
        funding_rate = sum(l["amount"] * l["rate"] for l in loans) / total_loan
    else:
        funding_rate = 0

    if total_invest > 0:
        invest_yield = sum(i["amount"] * i["rate"] for i in investments) / total_invest
    else:
        invest_yield = 0

    net_spread = round(invest_yield - funding_rate, 2)

    # Use Excel's Net P&L if available, otherwise compute
    excel_netpl = _num(_cell(ws, 5, COL_SUMMARY_INTEREST))
    if excel_netpl != 0:
        net_pl = excel_netpl
    else:
        net_pl = total_invest * invest_yield / 100 - total_loan * funding_rate / 100

    # Net P&L lũy kế from F6
    net_pl_cumulative = _num(_cell(ws, 6, COL_SUMMARY_INTEREST))

    # Investment breakdown
    td_amount = sum(i["amount"] for i in investments if i["type"] == "TD")
    bond_amount = sum(i["amount"] for i in investments if i["type"] == "BOND")
    td_pct = round(td_amount / total_invest * 100, 1) if total_invest > 0 else 0
    bond_pct = round(bond_amount / total_invest * 100, 1) if total_invest > 0 else 0

    total_han_muc = sum(lc["hanMuc"] for lc in limit_controls)

    # TSDB from Excel
    tsdb = _num(_cell(ws, 3, COL_SUMMARY_VALUE))

    summary = {
        "totalLoan": total_loan,
        "totalInvest": total_invest,
        "fundingRate": round(funding_rate, 2),
        "investYield": round(invest_yield, 2),
        "netSpread": net_spread,
        "netPL": net_pl,
        "netPLCumulative": net_pl_cumulative,
        "tdPct": td_pct,
        "bondPct": bond_pct,
        "tdAmount": td_amount,
        "bondAmount": bond_amount,
        "totalTSDB": tsdb,
        "totalHanMuc": total_han_muc,
    }

    # ---- 4. Duration Mismatches ----
    duration_mismatches = []
    for loan, inv in loan_invest_pairs:
        try:
            loan_end = datetime.strptime(loan["endDate"], "%Y-%m-%d")
            inv_end = datetime.strptime(inv["endDate"], "%Y-%m-%d")
            days_diff = abs((inv_end - loan_end).days)
            if days_diff > 0:
                duration_mismatches.append({
                    "investBank": inv["bank"],
                    "loanBank": loan["bank"],
                    "investEnd": inv["endDate"],
                    "loanEnd": loan["endDate"],
                    "daysDiff": days_diff,
                    "investAmt": inv["amount"],
                    "loanAmt": loan["amount"],
                })
        except ValueError:
            pass

    duration_mismatches.sort(key=lambda x: x["daysDiff"], reverse=True)

    # ---- 5. Aggregates by bank ----
    loan_by_bank = {}
    for l in loans:
        bank = l["bank"]
        if bank not in loan_by_bank:
            loan_by_bank[bank] = {"bank": bank, "total": 0, "count": 0, "avgRate": 0, "rates": []}
        loan_by_bank[bank]["total"] += l["amount"]
        loan_by_bank[bank]["count"] += 1
        loan_by_bank[bank]["rates"].append((l["amount"], l["rate"]))

    loan_by_bank_list = []
    for bank, data in loan_by_bank.items():
        avg_rate = sum(a * r for a, r in data["rates"]) / data["total"] if data["total"] > 0 else 0
        loan_by_bank_list.append({
            "bank": bank,
            "total": data["total"],
            "count": data["count"],
            "avgRate": round(avg_rate, 2),
        })

    invest_by_bank = {}
    for i in investments:
        bank = i["bank"]
        if bank not in invest_by_bank:
            invest_by_bank[bank] = {"bank": bank, "total": 0, "count": 0, "avgRate": 0, "rates": []}
        invest_by_bank[bank]["total"] += i["amount"]
        invest_by_bank[bank]["count"] += 1
        invest_by_bank[bank]["rates"].append((i["amount"], i["rate"]))

    invest_by_bank_list = []
    for bank, data in invest_by_bank.items():
        avg_rate = sum(a * r for a, r in data["rates"]) / data["total"] if data["total"] > 0 else 0
        invest_by_bank_list.append({
            "bank": bank,
            "total": data["total"],
            "count": data["count"],
            "avgRate": round(avg_rate, 2),
        })

    # ---- Build final dict ----
    financial_data = {
        "summary": summary,
        "limitControls": limit_controls,
        "loans": loans,
        "investments": investments,
        "durationMismatches": duration_mismatches,
        "loanByBank": loan_by_bank_list,
        "investByBank": invest_by_bank_list,
    }

    wb.close()
    return financial_data


# ---------------------------------------------------------------------------
# Output Generator
# ---------------------------------------------------------------------------

def write_treasury_data(data: dict, output_path: str):
    """Write FINANCIAL_DATA dict to treasury_data.py."""
    json_str = json.dumps(data, ensure_ascii=False, indent=4)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    content = f'''"""
Treasury Portfolio Financial Data.

AUTO-GENERATED by tools/sync_excel.py — DO NOT EDIT MANUALLY.
Last synced: {now}
Source: {EXCEL_FILENAME}
"""

FINANCIAL_DATA = {json_str}
'''

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"[SYNC] Written: {output_path}")


def git_push(project_dir: str):
    """Commit and push changes to trigger Vercel auto-deploy."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    try:
        subprocess.run(["git", "add", "lib/treasury_data.py"], cwd=project_dir, check=True)
        result = subprocess.run(
            ["git", "diff", "--cached", "--quiet"],
            cwd=project_dir,
        )
        if result.returncode == 0:
            print("[SYNC] No changes detected — skip push")
            return False

        subprocess.run(
            ["git", "commit", "-m", f"data: sync treasury data from Excel ({now})"],
            cwd=project_dir, check=True,
        )
        subprocess.run(["git", "push", "origin", "main"], cwd=project_dir, check=True)
        print("[SYNC] Pushed to GitHub — Vercel will auto-deploy!")
        return True
    except subprocess.CalledProcessError as e:
        print(f"[SYNC] Git error: {e}")
        return False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not os.path.isfile(EXCEL_PATH):
        print(f"[SYNC] ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    data = parse_excel(EXCEL_PATH)

    # Print summary
    s = data["summary"]
    print(f"\n{'='*50}")
    print(f"  Total Loans:       {s['totalLoan']/1e9:,.1f} ty ({len(data['loans'])} entries)")
    print(f"  Total Investments: {s['totalInvest']/1e9:,.1f} ty ({len(data['investments'])} entries)")
    print(f"  Funding Rate:      {s['fundingRate']:.2f}%")
    print(f"  Invest Yield:      {s['investYield']:.2f}%")
    print(f"  Net Spread:        {s['netSpread']:.2f}%")
    print(f"  Net P&L:           {s['netPL']/1e9:,.1f} ty")
    print(f"  Limit Controls:    {len(data['limitControls'])} banks")
    print(f"  Mismatches:        {len(data['durationMismatches'])} pairs")
    print(f"{'='*50}\n")

    write_treasury_data(data, OUTPUT_FILE)

    if "--push" in sys.argv:
        git_push(PROJECT_DIR)
    else:
        print("[SYNC] Run with --push to auto-deploy to Vercel")


if __name__ == "__main__":
    main()
