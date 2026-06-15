# -*- coding: utf-8 -*-
import openpyxl
import os
import json

base = os.path.join(os.path.expanduser("~"), "OneDrive - Công Ty Cổ Phần Dịch Vụ Giao Hàng Nhanh")
filepath = os.path.join(base, "Long Hoàng's files - Weekly Cash Report", "Working cap - Investment mapping", "Mapping WC-Invest 26.04.xlsx")

wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb["Mapping"]

print(f"=== FULL Mapping Sheet ({ws.max_row} rows x {ws.max_column} cols) ===\n")

for row_idx in range(1, min(ws.max_row + 1, 52)):
    vals = []
    for col_idx in range(1, min(ws.max_column + 1, 33)):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if v is None:
            vals.append("")
        elif isinstance(v, float):
            vals.append(f"{v:.4f}" if abs(v) < 1 else f"{v:,.0f}")
        else:
            vals.append(str(v)[:40])
    # Only print rows that have some content
    if any(v for v in vals):
        print(f"Row {row_idx:2d}: {vals}")
